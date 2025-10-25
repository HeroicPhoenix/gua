# -*- coding: utf-8 -*-
import os
import re
import time
import hashlib
from datetime import datetime
import pandas as pd


# ===== 统一列顺序（不存在文件时按此建表；存在则按列名对齐）=====
COL_ORDER = [
    "序号", "excel写入时间", "哈希值",
    "公历-年", "公历-月", "公历-日", "公历-时", "公历-星期",
    "农历-年", "农历-月", "农历-日", "农历-时",
    "干支-年", "干支-月", "干支-日", "干支-时",
    "旬空-年", "旬空-月", "旬空-日", "旬空-时",
    "时间1", "时间2",
    "月卦身", "世身", "八节", "神煞",
    "卦象文本", "卦象文本简介",
    "卦象名字", "本卦简称", "变卦简称",
]

# ===== 工具 =====
def md5_of_text(t: str) -> str:
    return hashlib.md5(t.encode("utf-8")).hexdigest()

def _weekday_to_num(w: str) -> str:
    if not w:
        return ""
    m = re.search(r"(星期|周)\s*([一二三四五六日天])", w)
    if not m:
        return ""
    ch = m.group(2)
    mapping = {"一":"1","二":"2","三":"3","四":"4","五":"5","六":"6","日":"7","天":"7"}
    return mapping.get(ch, "")

_CN_DIGITS = {"〇":0,"零":0,"一":1,"二":2,"三":3,"四":4,"五":5,"六":6,"七":7,"八":8,"九":9,"十":10,"廿":20,"卅":30,"初":0,"正":1}
def _cn_num_to_int(s: str) -> int:
    """将中文数字（含 初/十/廿/卅/正）转为整数，覆盖 1~31 常见组合。"""
    if s is None: return 0
    s = s.strip()
    s = "".join(ch for ch in s if ch in _CN_DIGITS)
    if not s:
        return 0
    if s in ("十",): return 10
    if s in ("廿",): return 20
    if s in ("卅",): return 30
    if s.startswith("初") and len(s) >= 2:
        if s[1] == "十":
            return 10
        return _CN_DIGITS.get(s[1], 0)
    if s.startswith("廿"):
        return 20 + _CN_DIGITS.get(s[1], 0) if len(s) > 1 else 20
    if s.startswith("卅"):
        return 30 + _CN_DIGITS.get(s[1], 0) if len(s) > 1 else 30
    if "十" in s:
        parts = s.split("十")
        if parts[0] == "":
            return 10 + _CN_DIGITS.get(parts[1], 0) if len(parts) > 1 else 10
        tens = _CN_DIGITS.get(parts[0], 0)
        ones = _CN_DIGITS.get(parts[1], 0) if len(parts) > 1 else 0
        return tens * 10 + ones
    return _CN_DIGITS.get(s, 0)

def _strip_paren(text: str) -> str:
    return re.sub(r"[\(（].*?[\)）]", "", text).strip()

# ===== 拆行，得到 5 条关键行（公历/农历/干支/旬空/节气行）=====
def _parse_first_block(full_text: str) -> dict:
    lines = [ln.rstrip() for ln in full_text.replace("\r\n", "\n").split("\n")]
    non_empty = [ln for ln in lines if ln.strip()]
    if len(non_empty) < 5:
        raise ValueError("文本格式不完整：关键行不足5行")
    line_g  = non_empty[0]  # 公历行（含“公历：”）
    line_n  = non_empty[1]  # 农历行（含“农历：”）
    line_gz = non_empty[2]  # 干支行
    line_xk = non_empty[3]  # 旬空行
    line_t  = non_empty[4]  # 节气时间行（“寒露…  霜降…”）
    return {"gl": line_g, "nl": line_n, "gz": line_gz, "xk": line_xk, "tline": line_t}

# ===== 逐项解析 =====
def _parse_gl(gl_line: str):
    gl = re.sub(r"^\s*公历\s*[：:]\s*", "", gl_line).strip()
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日\s*([0-2]?\d)[:：](\d{2})\s*(.*)$", gl)
    y = mo = d = hhmm = wk = ""
    if m:
        y = m.group(1)
        mo = m.group(2)
        d = m.group(3)
        hhmm = f"{int(m.group(4)):02d}:{m.group(5)}"
        wk = _weekday_to_num(m.group(6))
    else:
        m2 = re.search(r"([0-2]?\d)[:：](\d{2})", gl)
        if m2:
            hhmm = f"{int(m2.group(1)):02d}:{m2.group(2)}"
    return y, mo, d, hhmm, wk

def _parse_nl(nl_line: str):
    nl = re.sub(r"^\s*农历\s*[：:]\s*", "", nl_line).strip()
    m_year = re.search(r"^(.*?)年", nl)
    nl_year = _strip_paren(m_year.group(1)) if m_year else ""
    m_month = re.search(r"年(.*?)月", nl)
    nl_month = _cn_num_to_int(m_month.group(1) if m_month else "")
    m_day = re.search(r"月(.*?)(?:\s|$)", nl)
    day_raw = m_day.group(1) if m_day else ""
    day_raw = re.sub(r"^[大小]", "", day_raw)
    nl_day = _cn_num_to_int(day_raw)
    toks = nl.split()
    nl_time = toks[-1].strip() if toks else ""
    return nl_year, nl_month, nl_day, nl_time

def _parse_gz(gz_line: str):
    gz = re.sub(r"^\s*干支\s*[：:]\s*", "", gz_line).strip()
    parts = [p for p in re.split(r"\s+", gz) if p]
    while len(parts) < 4: parts.append("")
    return parts[0], parts[1], parts[2], parts[3]

def _parse_xk(xk_line: str):
    xk = re.sub(r"^\s*旬空\s*[：:]\s*", "", xk_line).strip()
    parts = [p for p in re.split(r"\s+", xk) if p]
    while len(parts) < 4: parts.append("")
    return parts[0], parts[1], parts[2], parts[3]

def _parse_time_line(tline: str):
    """
    节气时间行，例如：'寒露10月8日8:42  霜降10月23日11:52'
    要求：用 1 个以上空格分割；两侧 strip 去空格
    """
    if not tline:
        return "", ""
    parts = re.split(r"\s{1,}", tline.strip())
    # 某些界面会是两个以上空格，这个正则同样适配
    if len(parts) >= 2:
        return parts[0].strip(), parts[1].strip()
    if len(parts) == 1:
        return parts[0].strip(), ""
    return "", ""

def _parse_intro(intro_text: str):
    if not intro_text:
        return {"month_pos":"", "shi_shen":"", "ba_jie":"", "shen_sha":"", "name":"", "ben_gua":"", "bian_gua":"" , "intro_all":""}
    intro_all = intro_text.strip()
    lines = intro_all.replace("\r\n","\n").split("\n")
    first = lines[0] if lines else ""
    second = lines[1] if len(lines) > 1 else ""
    parts = [p.strip() for p in first.split("；") if p.strip()]
    name = parts[0] if len(parts) >= 1 else ""
    yueguashen = parts[1] if len(parts) >= 2 else ""
    shishen = parts[2] if len(parts) >= 3 else ""
    bajie = parts[3] if len(parts) >= 4 else ""
    yueguashen = re.sub(r"^月卦身", "", yueguashen).strip()
    shishen = re.sub(r"^世身在", "", shishen).strip()
    shensha = second.strip()
    shensha = re.sub(r"^\s*神煞\s*[：:]\s*", "", shensha)
    ben = bian = ""
    if "之" in name:
        idx = name.find("之")
        ben = name[:idx]
        bian = name[idx+1:]
    return {
        "month_pos": yueguashen, "shi_shen": shishen, "ba_jie": bajie,
        "shen_sha": shensha, "name": name, "ben_gua": ben, "bian_gua": bian,
        "intro_all": intro_all
    }

# ===== 生成行数据 =====
def build_excel_row(full_text: str, intro_text: str, write_dt: datetime) -> dict:
    """
    解析逻辑：
    1) 先用【原始】full_text 的第一行（公历行）做完整解析，得到 D~H（公历-年/月/日/时/星期）等字段；
    2) 再把“卦象文本”的第一行替换为基于 write_dt 的行：
       格式：'公历： YYYY年M月D日HH:MM 星期X'
    注意：仅替换 row['卦象文本'] 的首行；表内公历各字段仍是步骤1的解析结果。
    """
    # === 1) 先按原始文本完成解析（保持你现有的行为） ===
    blk = _parse_first_block(full_text)
    gl_y, gl_m, gl_d, gl_hhmm, gl_w = _parse_gl(blk["gl"])
    nl_y, nl_m, nl_d, nl_t = _parse_nl(blk["nl"])
    gz_y, gz_m, gz_d, gz_t = _parse_gz(blk["gz"])
    xk_y, xk_m, xk_d, xk_t = _parse_xk(blk["xk"])
    time1, time2 = _parse_time_line(blk.get("tline", ""))

    intro = _parse_intro(intro_text or "")

    # === 2) 仅替换“卦象文本”的第一行，使用 write_dt 生成标准行 ===
    def _weekday_cn(dt: datetime) -> str:
        # Python: Monday=0 ... Sunday=6
        mapping = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
        # 若你的本地以周一为0，这里正好对齐
        return mapping[dt.weekday()] if 0 <= dt.weekday() <= 6 else ""

    def _format_gl_line(dt: datetime) -> str:
        # 严格保持：'公历： YYYY年M月D日HH:MM 星期X'（冒号为中文全角，后面一个空格）
        return f"公历： {dt.year}年{dt.month}月{dt.day}日{dt:%H:%M} {_weekday_cn(dt)}"

    # 把 full_text 的第一条“非空行”替换为写入时间
    lines = full_text.replace("\r\n", "\n").split("\n")
    fmt_line = _format_gl_line(write_dt)
    replaced = False
    for i, ln in enumerate(lines):
        if ln.strip():
            lines[i] = fmt_line
            replaced = True
            break
    if not replaced:
        # 如果文本是空的或找不到非空行，就在开头补一行
        lines = [fmt_line] + lines
    patched_text = "\n".join(lines)

    # === 3) 组装 row：公历等字段用“原解析”，卦象文本用“替换后文本” ===
    row = {
        "序号": "",
        "excel写入时间": write_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "哈希值": md5_of_text(full_text),  # 仍以原始文本做哈希，避免频繁变化

        "公历-年": gl_y, "公历-月": gl_m, "公历-日": gl_d, "公历-时": gl_hhmm, "公历-星期": gl_w,
        "农历-年": nl_y, "农历-月": nl_m, "农历-日": nl_d, "农历-时": nl_t,

        "干支-年": gz_y, "干支-月": gz_m, "干支-日": gz_d, "干支-时": gz_t,
        "旬空-年": xk_y, "旬空-月": xk_m, "旬空-日": xk_d, "旬空-时": xk_t,

        "时间1": time1, "时间2": time2,

        "月卦身": intro["month_pos"], "世身": intro["shi_shen"], "八节": intro["ba_jie"], "神煞": intro["shen_sha"],
        "卦象文本": patched_text,            # ← 只在这里使用“Excel 插入时间”替换首行
        "卦象文本简介": intro["intro_all"],
        "卦象名字": intro["name"], "本卦简称": intro["ben_gua"], "变卦简称": intro["bian_gua"],
    }
    return row


# ===== 保存行到 Excel（按列名匹配；不存在则创建）=====
def save_row_to_excel(row: dict, path: str, extra_params=None):
    """
    仅写入“本次要写的列”，按列名匹配，不动其它列与格式/列宽；
    写入位置为：'excel写入时间' 列的“第一个空行”（从第2行开始）。
    若找不到该列，则自动建表头，并从第2行开始写。
    去重逻辑：与“最后一条已写入（excel写入时间非空）”的哈希值相同则跳过。
    """
    from openpyxl import load_workbook, Workbook
    import os

    extra_params = extra_params or []
    param_cols = [f"参数{i}" for i in range(1, len(extra_params) + 1)]

    # 1) 打开或新建工作簿
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        # 读取表头（第一行），保留原顺序
        headers = []
        if ws.max_row >= 1:
            for c in ws[1]:
                headers.append(str(c.value).strip() if c.value is not None else "")
        # 若第一行完全空，视为无表头
        if all(h == "" for h in headers):
            headers = []
    else:
        wb = Workbook()
        ws = wb.active
        headers = []

    # 2) 若没有表头，则以既定 COL_ORDER 起一个头（不强行加“参数列”，参数列按需追加在末尾）
    if not headers:
        headers = list(COL_ORDER)
        for j, name in enumerate(headers, 1):
            ws.cell(row=1, column=j, value=name)

    # 3) 需要写入的这些列：来自 row 的键 + 动态“参数1..N”
    #    仅确保这些列存在；缺失的列“只在末尾追加”，不改变已有列顺序/格式。
    def ensure_col(col_name: str):
        nonlocal headers
        if col_name not in headers:
            headers.append(col_name)
            ws.cell(row=1, column=len(headers), value=col_name)

    for k in row.keys():
        ensure_col(k)
    for c in param_cols:
        ensure_col(c)

    # 计算常用列索引（1-based）
    def col_idx(col_name: str) -> int:
        return headers.index(col_name) + 1 if col_name in headers else -1

    col_excel_time = col_idx("excel写入时间")
    col_hash = col_idx("哈希值")
    col_seq = col_idx("序号")

    # 4) 去重：与“最后一条已写入（excel写入时间非空）”记录的哈希相同则跳过
    #    从底向上找第一行 excel写入时间 非空的记录
    last_filled_hash = None
    if col_hash > 0 and col_excel_time > 0 and ws.max_row >= 2:
        for r in range(ws.max_row, 1, -1):
            t = ws.cell(row=r, column=col_excel_time).value
            if t is not None and str(t).strip() != "":
                last_filled_hash = ws.cell(row=r, column=col_hash).value
                break
    new_hash = str(row.get("哈希值", "")).strip()
    if last_filled_hash and new_hash and str(last_filled_hash).strip() == new_hash:
        print("（与上一条内容相同，跳过写入）")
        wb.save(path)  # 以防刚才有补表头/补列
        return False

    # 5) 计算写入的“目标行号”：从第2行开始，找“excel写入时间”列的第一个空行
    #    若没有该列（极端情况），则直接用 ws.max_row+1
    if col_excel_time > 0:
        write_row = None
        last_seq_above = None
        for r in range(2, ws.max_row + 2):  # +2 允许在末尾新开一行
            cell_val = ws.cell(row=r, column=col_excel_time).value
            if cell_val is None or str(cell_val).strip() == "":
                write_row = r
                # 同时记录“上一条”的序号，便于递增
                if r > 2 and col_seq > 0:
                    last_seq_above = ws.cell(row=r - 1, column=col_seq).value
                break
        if write_row is None:
            write_row = ws.max_row + 1
            if col_seq > 0 and ws.max_row >= 2:
                last_seq_above = ws.cell(row=ws.max_row, column=col_seq).value
    else:
        write_row = ws.max_row + 1
        last_seq_above = ws.cell(row=ws.max_row, column=col_seq).value if (col_seq > 0 and ws.max_row >= 2) else None

    # 6) 自动写入“序号”（当且仅当存在该列）：延续上一条的序号+1；若不可解析则用“去掉表头的行号”
    if col_seq > 0:
        try:
            next_seq = int(last_seq_above) + 1 if last_seq_above not in (None, "") else (write_row - 1)
        except Exception:
            next_seq = (write_row - 1)
        row["序号"] = next_seq

    # 7) 组装要写入的键值：仅对“本次要写的列”赋值，其它列完全不动（包括你手工加的列/手工写的数据）
    values = {k: ("" if v is None else v) for k, v in row.items()}
    for i, p in enumerate(extra_params, 1):
        values[f"参数{i}"] = "" if p is None else str(p)

    # 8) 真正写入：仅写需要的列（按列名找位置）；其它列一个字节不碰，格式/列宽保持
    for key, v in values.items():
        j = col_idx(key)
        if j > 0:
            ws.cell(row=write_row, column=j, value=v)

    # 9) 保存
    wb.save(path)
    print(f"已写入：{path}（第 {write_row} 行）")
    return True



# ====== Excel 中批量插入空列（UI 按钮会调用）=====
def insert_blank_cols(path: str, x_col_1based: int, col_names: list):
    if not isinstance(x_col_1based, int) or x_col_1based < 1:
        raise ValueError("X 必须是从 1 开始的正整数。")
    if (not isinstance(col_names, list)
        or any(not isinstance(n, str) or n.strip() == "" for n in col_names)):
        raise ValueError("列名必须为非空字符串列表。")

    if os.path.exists(path):
        df = pd.read_excel(path)
    else:
        df = pd.DataFrame()

    insert_loc = max(0, min(x_col_1based - 1, len(df.columns)))
    for i, nm in enumerate(col_names):
        df.insert(loc=insert_loc + i, column=nm, value="")

    df.to_excel(path, index=False)
    print(f"已在第 {x_col_1based} 列前插入 {len(col_names)} 列：{col_names}")
